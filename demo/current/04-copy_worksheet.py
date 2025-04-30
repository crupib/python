import os
import openpyxl
def copy_py_reports(source_workbook: str, destination_workbook: str):
    source_wb = openpyxl.load_workbook(os.path.abspath(source_workbook))
    dest_wb = openpyxl.load_workbook(os.path.abspath(destination_workbook))

    all_sheets = source_wb.sheetnames

    for sheet in all_sheets:
        source_ws = source_wb[sheet]
        dest_ws = dest_wb.create_sheet(title="mysheetname")

        # Copy the cell values and formatting from the source worksheet to the destination worksheet
        for row in source_ws.iter_rows(min_row=1, 
                                       max_col=source_ws.max_column,
                                       max_row=source_ws.max_row):
                for cell in row:
                    dest_cell = dest_ws.cell(row=cell.row, column=cell.column, 
                                             value=cell.value)
                    dest_cell.number_format = cell.number_format
                    dest_cell.font = copy(cell.font)
                    dest_cell.border = copy(cell.border)
                    dest_cell.fill = copy(cell.fill)
                    dest_cell.alignment = copy(cell.alignment)

                    # Check if the source cell is part of a merged cell range
                    if cell.coordinate in source_ws.merged_cells:
                        # Get the merged cell range and merge the corresponding cells
                        # in the destination worksheet
                        for range_ in source_ws.merged_cells.ranges:
                            if cell.coordinate in range_:
                                dest_ws.merge_cells(start_row=range_.min_row, 
                                                    start_column=range_.min_col,
                                                    end_row=range_.max_row, 
                                                    end_column=range_.max_col)
                                break

    dest_wb.save(filename=destination_workbook)
copy_py_reports("myfile.xlsx","myfile_saved.xlsx")
