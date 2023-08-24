import openpyxl
from openpyxl.styles import PatternFill

workbook = openpyxl.load_workbook("myfile.xlsx")
# adds a new worksheet to an existing workbook
worksheet = workbook.create_sheet(title='mysheet')

# fill grey color
grey_fill = PatternFill(start_color='CECACA',
                        end_color='CECACA',
                        fill_type='solid')

# fill red color
red_fill = PatternFill(start_color='990000',
                       end_color='990000',
                       fill_type='solid')

worksheet['A1'].fill = grey_fill
worksheet['A2'].fill = red_fill
workbook.save(filename="myfile.xlsx")
